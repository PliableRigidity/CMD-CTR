"""Phase 12C hardware import pipeline — BOM and inventory ingestion."""
from __future__ import annotations

import csv
import json
import os
import re
from pathlib import Path
from typing import Iterable

from backend.app.services.desktop_service import DesktopService
from backend.app.services.hardware_service import HardwareService


_QTY_KEYS = {"qty", "quantity", "count", "qnty", "qty per board", "quantity per pcb"}
_VALUE_KEYS = {"value", "val", "component value"}
_NAME_KEYS = {"name", "part", "part name", "component", "description", "desc"}
_FOOTPRINT_KEYS = {"footprint", "package", "pcb footprint"}
_MANUFACTURER_KEYS = {"manufacturer", "mfg", "mfr"}
_MPN_KEYS = {"part number", "part_number", "mpn", "mfg part number", "manufacturer part number", "pn"}
_REF_KEYS = {"reference", "references", "ref", "refs", "designator", "designators"}


def _key(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", " ", (value or "").lower()).strip()


def _first(row: dict, keys: set[str]) -> str:
    for original, value in row.items():
        if _key(original) in keys and value is not None:
            return str(value).strip()
    return ""


def _quantity(row: dict) -> int:
    raw = _first(row, _QTY_KEYS)
    if not raw:
        refs = _first(row, _REF_KEYS)
        if refs:
            return max(1, len([r for r in re.split(r"[\s,;]+", refs) if r.strip()]))
        return 1
    try:
        return max(0, int(float(str(raw).strip())))
    except ValueError:
        return 1


def _project_from_path(path: Path) -> str:
    stem = path.stem
    stem = re.sub(r"(?i)(?:^bom[_\-\s]*|[_\-\s]*(?:bom|ibom|components?|parts?|inventory|kicad)$)", "", stem)
    stem = re.sub(r"[_\-]+", " ", stem).strip()
    return stem or path.parent.name


def _read_csv(path: Path) -> list[dict]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        sample = handle.read(4096)
        handle.seek(0)
        try:
            dialect = csv.Sniffer().sniff(sample)
        except csv.Error:
            dialect = csv.excel
        return [dict(row) for row in csv.DictReader(handle, dialect=dialect)]


def _read_xlsx(path: Path) -> list[dict]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("Excel import requires openpyxl to be installed.") from exc
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(cell or "").strip() for cell in rows[0]]
    output = []
    for values in rows[1:]:
        if not any(values):
            continue
        output.append({headers[i] if i < len(headers) else f"column_{i}": values[i] for i in range(len(values))})
    return output


def _extract_kicad_property(block: str, name: str) -> str:
    match = re.search(
        rf'\(property\s+"{re.escape(name)}"\s+"((?:\\"|[^"])*)"',
        block,
        re.I,
    )
    return match.group(1).replace('\\"', '"').strip() if match else ""


def _read_kicad_schematic(path: Path) -> list[dict]:
    text = path.read_text(encoding="utf-8", errors="ignore")
    rows = []
    for block in re.split(r"\(symbol\s+", text):
        reference = _extract_kicad_property(block, "Reference")
        value = _extract_kicad_property(block, "Value")
        footprint = _extract_kicad_property(block, "Footprint")
        if not reference or not value:
            continue
        if reference.startswith("#") or value.lower() in {"power", "gnd", "vcc"}:
            continue
        rows.append({
            "Reference": reference,
            "Value": value,
            "Qty": 1,
            "Footprint": footprint,
        })
    merged: dict[tuple[str, str], dict] = {}
    for row in rows:
        key = (row["Value"], row.get("Footprint", ""))
        if key in merged:
            merged[key]["Qty"] += 1
            merged[key]["Reference"] += f",{row['Reference']}"
        else:
            merged[key] = row
    return list(merged.values())


def _read_kicad_project(path: Path) -> list[dict]:
    try:
        json.loads(path.read_text(encoding="utf-8", errors="ignore") or "{}")
    except json.JSONDecodeError:
        pass
    rows = []
    for schematic in sorted(path.parent.glob("*.kicad_sch")):
        rows.extend(_read_kicad_schematic(schematic))
    if not rows:
        raise ValueError(f"No schematic components found beside {path.name}. Export a KiCad BOM CSV for best results.")
    return rows


def _read_markdown_components(path: Path) -> list[dict]:
    rows = []
    for line in path.read_text(encoding="utf-8", errors="ignore").splitlines():
        stripped = line.strip()
        table_match = re.match(r"^\|\s*(?P<name>[^|]+?)\s*\|\s*(?P<qty>\d+)\s*\|", stripped)
        list_match = re.match(r"^[-*]\s*(?:(?P<qty>\d+)\s*[x×]\s*)?(?P<name>[A-Za-z0-9][A-Za-z0-9_.\- ]{1,80})", stripped)
        match = table_match or list_match
        if not match:
            continue
        name = match.group("name").strip()
        if name.lower() in {"part", "name", "component"}:
            continue
        rows.append({"Part": name, "Qty": int(match.group("qty") or 1)})
    if not rows:
        raise ValueError(f"No Markdown component list rows found in {path.name}.")
    return rows


def read_tabular_file(path: str) -> list[dict]:
    source = Path(path)
    suffix = source.suffix.lower()
    if suffix in {".csv", ".tsv"}:
        return _read_csv(source)
    if suffix in {".xlsx", ".xlsm"}:
        return _read_xlsx(source)
    if suffix == ".kicad_sch":
        return _read_kicad_schematic(source)
    if suffix == ".kicad_pro":
        return _read_kicad_project(source)
    if suffix in {".md", ".markdown", ".txt"}:
        return _read_markdown_components(source)
    raise ValueError(f"Unsupported import file type: {source.suffix}")


def resolve_import_path(path_or_name: str) -> str:
    candidate = Path(path_or_name).expanduser()
    if candidate.is_file():
        return str(candidate)
    svc = DesktopService()
    suffix = candidate.suffix.lower().lstrip(".")
    query = candidate.name if candidate.name else path_or_name
    results = svc.search_files(query=query, extension=suffix, max_results=10)
    if not results:
        raise FileNotFoundError(f"No file found for '{path_or_name}' in indexed trusted locations.")
    exact = [item for item in results if item["name"].lower() == query.lower()]
    if len(exact) == 1:
        return exact[0]["path"]
    if len(results) == 1:
        return results[0]["path"]
    names = ", ".join(item["path"] for item in results[:5])
    raise RuntimeError(f"Multiple files matched '{path_or_name}': {names}")


def _component_from_row(row: dict) -> dict:
    value = _first(row, _VALUE_KEYS)
    part_number = _first(row, _MPN_KEYS)
    description = _first(row, _NAME_KEYS)
    manufacturer = _first(row, _MANUFACTURER_KEYS)
    footprint = _first(row, _FOOTPRINT_KEYS)
    name = part_number or value or description
    if not name:
        name = _first(row, _REF_KEYS)
    return {
        "raw_name": name,
        "name": name,
        "quantity": _quantity(row),
        "value": value,
        "footprint": footprint,
        "manufacturer": manufacturer,
        "part_number": part_number,
        "raw": {str(k): "" if v is None else str(v) for k, v in row.items()},
    }


class HardwareImportService:
    def __init__(self, hardware: HardwareService | None = None) -> None:
        self.hardware = hardware or HardwareService()

    def import_file(
        self,
        path: str,
        *,
        source_type: str = "auto",
        project_name: str = "",
    ) -> dict:
        resolved = resolve_import_path(path)
        source = Path(resolved)
        rows = read_tabular_file(resolved)
        if source_type == "auto":
            lowered = source.name.lower()
            source_type = "bom" if (
                "bom" in lowered
                or project_name
                or source.suffix.lower() in {".kicad_pro", ".kicad_sch", ".md", ".markdown", ".txt"}
            ) else "inventory"
        if source_type not in {"bom", "inventory"}:
            raise ValueError("source_type must be 'bom', 'inventory', or 'auto'")

        project = None
        project_created = False
        if source_type == "bom":
            inferred = project_name.strip() or _project_from_path(source)
            project, project_created = self.hardware.get_or_create_project(
                inferred,
                notes=f"Created from BOM import: {source.name}",
            )

        errors: list[str] = []
        items: list[dict] = []
        created_parts = updated_parts = linked_parts = imported_count = 0
        merged: dict[tuple[str, str, str], dict] = {}
        for index, row in enumerate(rows, start=2):
            component = _component_from_row(row)
            if not component["name"]:
                errors.append(f"Row {index}: no part name/value/part number found")
                continue
            key = (
                re.sub(r"[^a-z0-9]+", "", component["name"].lower()),
                component["manufacturer"].lower(),
                component["part_number"].lower(),
            )
            if key in merged:
                merged[key]["quantity"] += component["quantity"]
            else:
                merged[key] = component

        for component in merged.values():
            part, action = self.hardware.upsert_import_part(
                name=component["name"],
                quantity=component["quantity"],
                source_type=source_type,
                manufacturer=component["manufacturer"],
                part_number=component["part_number"],
                value=component["value"],
                footprint=component["footprint"],
            )
            if action == "created":
                created_parts += 1
            else:
                updated_parts += 1
            if project:
                self.hardware.assign_part_to_project(
                    project["id"],
                    part["id"],
                    quantity_required=max(1, component["quantity"]),
                    notes=f"Imported from {source.name}",
                    is_required=1,
                    source="BOM import",
                )
                linked_parts += 1
            imported_count += 1
            items.append({
                **component,
                "part_id": part["id"],
                "action": action,
            })

        import_record = self.hardware.record_import(
            source_path=resolved,
            source_type=source_type,
            project=project,
            rows_total=len(rows),
            imported_count=imported_count,
            created_parts=created_parts,
            updated_parts=updated_parts,
            linked_parts=linked_parts,
            errors=errors,
            items=items,
        )
        readiness = self.hardware.get_build_readiness(project["id"]) if project else None
        return {
            "ok": True,
            "import": import_record,
            "project": project,
            "project_created": project_created,
            "readiness": readiness,
            "items": items,
            "errors": errors,
            "summary": {
                "source_path": resolved,
                "source_type": source_type,
                "rows_total": len(rows),
                "imported_count": imported_count,
                "created_parts": created_parts,
                "updated_parts": updated_parts,
                "linked_parts": linked_parts,
            },
        }
